import os
import re
import glob
from collections import defaultdict
import traceback
import PyPDF2
import pandas as pd
import matplotlib.pyplot as plt
import seaborn as sns
from datetime import datetime
import sys
import argparse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
import markdown
import numpy as np
import base64
import io
import platform
import json

# Set up Korean font for matplotlib
if platform.system() == "Darwin":  # macOS
    plt.rcParams["font.family"] = "AppleGothic"
elif platform.system() == "Windows":
    plt.rcParams["font.family"] = "Malgun Gothic"
else:  # Linux
    plt.rcParams["font.family"] = "NanumGothic"

plt.rcParams["axes.unicode_minus"] = False  # 마이너스 기호 깨짐 방지

CONFIG_FILE = "config.json"


def get_last_folder():
    """마지막으로 선택한 폴더 경로를 반환합니다."""
    try:
        if os.path.exists(CONFIG_FILE):
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                config = json.load(f)
                last_folder = config.get("last_folder")
                if last_folder and os.path.exists(last_folder):
                    return last_folder
    except Exception as e:
        print(f"설정 파일 로드 중 오류 발생: {e}")
    return None


def save_last_folder(folder_path):
    """선택한 폴더 경로를 저장합니다."""
    try:
        config = {"last_folder": folder_path}
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(config, f, ensure_ascii=False, indent=2)
    except Exception as e:
        print(f"설정 파일 저장 중 오류 발생: {e}")


def find_pdf_files(root_dir=".", log_callback=None):
    """Find all PDF files in all subdirectories without specific naming restrictions."""
    pdf_files = []

    try:
        # 검색 시작 디렉토리의 절대 경로를 구함
        root_abs = os.path.abspath(root_dir)

        if log_callback:
            log_callback(f"검색 시작 디렉토리: {root_abs}\n")

        for dirpath, dirnames, filenames in os.walk(root_abs):
            # 현재 디렉토리가 root_dir의 하위 디렉토리인지 확인
            rel_path = os.path.relpath(dirpath, root_abs)
            if rel_path == ".":
                path_parts = []
            else:
                # Windows에서도 올바르게 동작하도록 normpath 사용
                path_parts = os.path.normpath(rel_path).split(os.sep)

            # 언더스코어로 시작하는 디렉토리 건너뛰기
            should_skip = False
            for part in path_parts:
                if part.startswith("_"):
                    if log_callback:
                        log_callback(f"건너뛰는 디렉토리: {rel_path}\n")
                    should_skip = True
                    break

            if should_skip:
                dirnames[:] = []  # 하위 디렉토리 탐색 중지
                continue

            for filename in filenames:
                if filename.lower().endswith(".pdf"):
                    # 상대 경로 생성
                    if rel_path == ".":
                        rel_file_path = filename
                    else:
                        rel_file_path = os.path.join(rel_path, filename)

                    # 절대 경로로 변환
                    abs_file_path = os.path.join(root_abs, rel_file_path)
                    pdf_files.append(abs_file_path)

                    if log_callback:
                        log_callback(f"PDF 파일 발견: {abs_file_path}\n")

        return pdf_files
    except Exception as e:
        if log_callback:
            log_callback(f"PDF 파일 검색 중 오류 발생: {e}\n")
        return []


def extract_text_from_pdf(pdf_path, log_callback=None):
    """Extract text content from a PDF file."""
    try:
        # 파일 경로를 절대 경로로 변환
        abs_path = os.path.abspath(pdf_path)

        if not os.path.exists(abs_path):
            if log_callback:
                log_callback(f"파일을 찾을 수 없음: {abs_path}\n")
            return ""

        try:
            with open(abs_path, "rb") as file:
                reader = PyPDF2.PdfReader(file)
                text = ""
                for page in reader.pages:
                    text += page.extract_text()
            return text
        except PermissionError:
            if log_callback:
                log_callback(f"파일 접근 권한 없음: {abs_path}\n")
            return ""
        except Exception as e:
            if log_callback:
                log_callback(f"PDF 파일 읽기 오류 {abs_path}: {e}\n")
            return ""
    except Exception as e:
        if log_callback:
            log_callback(f"파일 처리 오류 {pdf_path}: {e}\n")
        return ""


def parse_pdf_content(text, pdf_path, log_callback=None):
    """Parse PDF text content with improved data extraction and error indication."""
    data = {}

    # Extract filename
    filename = os.path.basename(pdf_path)
    data["filename"] = filename

    # Extract region from filename if available (optional)
    region_match = re.search(r"\(([^)]+)\)", filename)
    if region_match:
        data["region"] = region_match.group(1)
    else:
        data["region"] = "Unknown"  # Set default value if pattern not found

    # Extract timestamp from filename if available (optional)
    timestamp_match = re.search(r"(\d{4}_\d{2}_\d{2}_\d{2}_\d{2}_\d{2})", filename)
    if timestamp_match:
        data["timestamp"] = timestamp_match.group(1).replace("_", "-")
    else:
        # Use file modification date as fallback
        try:
            mod_time = os.path.getmtime(pdf_path)
            data["timestamp"] = datetime.fromtimestamp(mod_time).strftime(
                "%Y-%m-%d-%H-%M-%S"
            )
        except:
            data["timestamp"] = "Unknown"

    # Debug information about the file being processed
    if log_callback:
        log_callback(f"\n==== 파일 처리 중: {filename} ====\n")

    # Playtime extraction - enhanced pattern matching
    playtime_found = False

    # List of pattern strategies to try
    playtime_patterns = [
        r"Play\s*Time\s*\n\s*(\d+\.?\d*)\s*s",  # Standard pattern: "Play Time" followed by newline and digits with 's'
        r"Play\s+Time\s*[\r\n\s]*(\d+\.?\d*)\s*s",  # More flexible with various whitespace
        r"Play Time.*?(\d+\.\d+)\s*s",  # Any content between "Play Time" and the number
        r"Play\s*Time[^0-9]*(\d+\.?\d+)",  # Anything except digits between "Play Time" and number
        r"Play\s*Time.*?(\d+\.?\d+)",  # Most general pattern
        r"play\s*time.*?(\d+\.?\d+)\s*s",  # Case insensitive variation
        r"duration.*?(\d+\.?\d+)\s*s",  # Alternative term "duration"
        r"length.*?(\d+\.?\d+)\s*s",  # Alternative term "length"
        r"time.*?(\d+\.?\d+)\s*s",  # Last resort - any "time" followed by a number and "s"
    ]

    # Try each pattern strategy
    for i, pattern in enumerate(playtime_patterns):
        playtime_match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if playtime_match:
            try:
                # Extract the value from the matched group
                data["playtime"] = float(playtime_match.group(1))
                if log_callback:
                    log_callback(f"패턴 {i+1}로 재생 시간 찾음: {data['playtime']} s\n")
                playtime_found = True
                break
            except (ValueError, IndexError) as e:
                if log_callback:
                    log_callback(f"패턴 {i+1}로 재생 시간 파싱 오류: {e}\n")

    # If still not found, try looking directly at play time context
    if not playtime_found:
        if log_callback:
            log_callback("재생 시간 패턴 실패, Play Time 섹션 컨텍스트 검사 중...\n")

        # Get 200 characters surrounding "Play Time" mention
        play_time_idx = text.find("Play Time")
        if play_time_idx == -1:
            play_time_idx = text.lower().find("play time")

        if play_time_idx != -1:
            context_start = max(0, play_time_idx - 20)
            context_end = min(len(text), play_time_idx + 180)
            context = text[context_start:context_end]
            if log_callback:
                log_callback(f"Play Time 컨텍스트: '{context}'\n")

            # Look for any number in this context
            number_match = re.search(r"(\d+\.?\d*)\s*s", context)
            if number_match:
                try:
                    data["playtime"] = float(number_match.group(1))
                    if log_callback:
                        log_callback(
                            f"컨텍스트에서 재생 시간 추출: {data['playtime']} s\n"
                        )
                    playtime_found = True
                except (ValueError, IndexError) as e:
                    if log_callback:
                        log_callback(f"컨텍스트에서 재생 시간 파싱 오류: {e}\n")

    # As a very last resort, search for any number followed by "s" or "seconds" in the first 10% of the document
    if not playtime_found:
        if log_callback:
            log_callback("문서에서 시간과 관련된 값 검색 중...\n")
        search_section = text[: int(len(text) * 0.1)]  # First 10% of document
        time_matches = re.findall(
            r"(\d+\.?\d*)\s*(?:s|sec|seconds)", search_section, re.IGNORECASE
        )
        if time_matches:
            try:
                # Use the first reasonable value found (between 1 and 3600 seconds)
                for potential_time in time_matches:
                    time_val = float(potential_time)
                    if 1 <= time_val <= 3600:
                        data["playtime"] = time_val
                        if log_callback:
                            log_callback(
                                f"가능한 재생 시간 값 발견: {data['playtime']} s\n"
                            )
                        playtime_found = True
                        break
            except (ValueError, IndexError) as e:
                if log_callback:
                    log_callback(f"가능한 시간 값 파싱 오류: {e}\n")

    # If still not found, set to -1 to indicate an error
    if not playtime_found:
        data["playtime"] = -1  # Error value
        if log_callback:
            log_callback("재생 시간을 찾을 수 없어 오류 값(-1) 설정\n")

    # Extract FPS information
    fps_patterns = [
        r"FPS\s*.*?\s*Avg\s*:\s*(\d+\.?\d*)",  # Standard pattern
        r"FPS.*?average.*?(\d+\.?\d*)",  # Alternative "average" wording
        r"Frames\s*Per\s*Second.*?(\d+\.?\d*)",  # Full "frames per second" term
        r"Frame\s*Rate.*?(\d+\.?\d*)",  # Alternative "frame rate" term
        r"Average\s*FPS.*?(\d+\.?\d*)",  # "Average FPS" pattern
        r"fps.*?(\d+\.?\d*)",  # Simple "fps" mention
    ]

    fps_found = False

    for i, pattern in enumerate(fps_patterns):
        fps_match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if fps_match:
            try:
                data["fps"] = float(fps_match.group(1))
                if log_callback:
                    log_callback(f"패턴 {i+1}로 FPS 찾음: {data['fps']}\n")
                fps_found = True
                break
            except (ValueError, IndexError) as e:
                if log_callback:
                    log_callback(f"패턴 {i+1}로 FPS 파싱 오류: {e}\n")

    # If not found, use error value (-1)
    if not fps_found:
        data["fps"] = -1  # Error value
        if log_callback:
            log_callback("FPS를 찾을 수 없어 오류 값(-1) 설정\n")

    # Bandwidth patterns
    bandwidth_patterns = [
        r"Bandwidth\s*.*?\s*Avg\s*:\s*(\d+\.?\d*)\s*Mbps",  # Standard pattern
        r"Bandwidth.*?(\d+\.?\d*)\s*Mbps",  # Simpler pattern
        r"Average\s*Bandwidth.*?(\d+\.?\d*)\s*Mbps",  # "Average Bandwidth" term
        r"Network\s*Speed.*?(\d+\.?\d*)\s*Mbps",  # Alternative "Network Speed" term
        r"Data\s*Rate.*?(\d+\.?\d*)\s*Mbps",  # Alternative "Data Rate" term
        r"(\d+\.?\d*)\s*Mbps",  # Last resort - any Mbps value
    ]

    bw_found = False

    for i, pattern in enumerate(bandwidth_patterns):
        bandwidth_match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if bandwidth_match:
            try:
                data["bandwidth"] = float(bandwidth_match.group(1))
                if log_callback:
                    log_callback(
                        f"패턴 {i+1}로 대역폭 찾음: {data['bandwidth']} Mbps\n"
                    )
                bw_found = True
                break
            except (ValueError, IndexError) as e:
                if log_callback:
                    log_callback(f"패턴 {i+1}로 대역폭 파싱 오류: {e}\n")

    # If not found, use error value (-1)
    if not bw_found:
        data["bandwidth"] = -1  # Error value
        if log_callback:
            log_callback("대역폭을 찾을 수 없어 오류 값(-1) 설정\n")

    # RTT patterns
    rtt_patterns = [
        r"Round Trip Time\s*.*?\s*Avg\s*:\s*(\d+\.?\d*)\s*ms",  # Standard pattern
        r"RTT.*?(\d+\.?\d*)\s*ms",  # RTT abbreviation
        r"Round\s*Trip.*?(\d+\.?\d*)\s*ms",  # Partial "Round Trip" term
        r"Latency.*?(\d+\.?\d*)\s*ms",  # Alternative "Latency" term
        r"Ping.*?(\d+\.?\d*)\s*ms",  # Alternative "Ping" term
        r"Response\s*Time.*?(\d+\.?\d*)\s*ms",  # Alternative "Response Time" term
    ]

    rtt_found = False

    for i, pattern in enumerate(rtt_patterns):
        rtt_match = re.search(pattern, text, re.DOTALL | re.IGNORECASE)
        if rtt_match:
            try:
                data["rtt"] = float(rtt_match.group(1))
                if log_callback:
                    log_callback(f"패턴 {i+1}로 RTT 찾음: {data['rtt']} ms\n")
                rtt_found = True
                break
            except (ValueError, IndexError) as e:
                if log_callback:
                    log_callback(f"패턴 {i+1}로 RTT 파싱 오류: {e}\n")

    # If not found, use error value (-1)
    if not rtt_found:
        data["rtt"] = -1  # Error value
        if log_callback:
            log_callback("RTT를 찾을 수 없어 오류 값(-1) 설정\n")

    if log_callback:
        log_callback(f"==== 파일 처리 완료: {filename} ====\n")
    return data


def get_folder_path(pdf_path):
    """Extract folder path from the full path."""
    try:
        # 경로를 정규화하고 디렉토리 부분만 추출
        norm_path = os.path.normpath(pdf_path)
        dir_path = os.path.dirname(norm_path)

        # 빈 문자열인 경우 현재 디렉토리를 의미
        return dir_path if dir_path else "."
    except Exception as e:
        print(f"Error getting folder path for {pdf_path}: {e}")
        return "."


def process_pdf_files(root_dir=".", log_callback=None, progress_callback=None):
    """Process all PDF files in the directory structure."""
    if log_callback:
        log_callback("PDF 파일 검색 중...\n")

    # 상위 폴더에 reports 폴더 생성
    root_abs = os.path.abspath(root_dir)
    parent_dir = os.path.dirname(root_abs)
    reports_dir = os.path.join(parent_dir, "reports")

    if log_callback:
        log_callback(f"선택한 폴더: {root_abs}\n")
        log_callback(f"상위 폴더: {parent_dir}\n")
        log_callback(f"리포트 폴더: {reports_dir}\n")

    try:
        os.makedirs(reports_dir, exist_ok=True)
        if log_callback:
            log_callback(f"리포트 저장 폴더 생성됨: {reports_dir}\n")
    except Exception as e:
        if log_callback:
            log_callback(f"리포트 폴더 생성 중 오류 발생: {e}\n")
            log_callback("현재 폴더에 reports 디렉토리 생성 시도...\n")

        # 현재 폴더에 reports 디렉토리 생성 시도
        reports_dir = os.path.join(root_abs, "reports")
        try:
            os.makedirs(reports_dir, exist_ok=True)
            if log_callback:
                log_callback(f"현재 폴더에 리포트 저장 폴더 생성됨: {reports_dir}\n")
        except Exception as e2:
            if log_callback:
                log_callback(f"현재 폴더에 리포트 폴더 생성 중 오류 발생: {e2}\n")
            return None

    # Initialize folder_data with reports_dir and root_abs information
    folder_data = defaultdict(lambda: defaultdict(list))
    folder_data["_config"] = {"reports_dir": reports_dir, "root_abs": root_abs}

    pdf_files = find_pdf_files(root_dir, log_callback)
    total = len(pdf_files)

    if not pdf_files:
        if log_callback:
            log_callback("처리할 PDF 파일을 찾을 수 없습니다.\n")
        return folder_data  # Return with reports_dir info even if no PDFs found

    if log_callback:
        log_callback(f"총 {len(pdf_files)}개의 PDF 파일을 찾았습니다.\n")
        for pdf_file in pdf_files:
            log_callback(f"  - {pdf_file}\n")
        log_callback("\n파일 처리 시작...\n")

    for idx, pdf_file in enumerate(pdf_files):
        if log_callback:
            log_callback(f"\n처리 중: {pdf_file}\n")

        text = extract_text_from_pdf(pdf_file, log_callback)
        text = re.sub(r"\s*\.\s*", ".", text)
        if not text:
            if log_callback:
                log_callback(
                    f"경고: {pdf_file} 파일에서 텍스트를 추출할 수 없습니다.\n"
                )
            continue

        data = parse_pdf_content(text, pdf_file, log_callback)
        if not data:
            if log_callback:
                log_callback(
                    f"경고: {pdf_file} 파일에서 데이터를 파싱할 수 없습니다.\n"
                )
            continue

        folder = get_folder_path(pdf_file)
        folder_data[folder]["files"].append(data)

        if log_callback:
            log_callback("추출된 데이터:\n")
            for key, value in data.items():
                if key != "filename":  # filename은 이미 위에서 표시했으므로 제외
                    log_callback(f"  - {key}: {value}\n")

        if progress_callback:
            progress_callback(idx + 1, total)

    if log_callback:
        log_callback("\nPDF 파일 처리가 완료되었습니다.\n")

    # 마지막에 folder_data와 내부를 모두 dict로 변환해서 반환
    def deep_dict(obj):
        if isinstance(obj, defaultdict):
            return {k: deep_dict(v) for k, v in obj.items()}
        return obj

    return deep_dict(folder_data)


def generate_folder_report(folder_data, log_callback=None):
    """Generate markdown and HTML reports for the folder data."""
    if log_callback:
        log_callback("마크다운 및 HTML 리포트 생성 중...\n")

    try:
        # Get reports directory and root_abs from folder_data
        reports_dir = folder_data.get("_config", {}).get("reports_dir")
        root_abs = folder_data.get("_config", {}).get("root_abs", "")
        if not reports_dir:
            if log_callback:
                log_callback("reports 디렉토리 경로를 찾을 수 없습니다.\n")
            return None, None, None, None

        if log_callback:
            log_callback(f"리포트 저장 위치: {reports_dir}\n")

        # 현재 시간을 이용한 타임스탬프 생성
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # 파일명 생성
        fixed_md = os.path.join(reports_dir, "folder_metrics_report.md")
        timestamped_md = os.path.join(
            reports_dir, f"folder_metrics_report_{timestamp}.md"
        )
        fixed_html = os.path.join(reports_dir, "folder_metrics_report.html")
        timestamped_html = os.path.join(
            reports_dir, f"folder_metrics_report_{timestamp}.html"
        )

        if log_callback:
            log_callback("리포트 파일 생성 중...\n")

        # HTML 스타일 정의
        html_style = """
        <style>
            body { font-family: Arial, sans-serif; margin: 40px; line-height: 1.6; }
            table { border-collapse: collapse; width: 100%; margin: 20px 0; }
            th, td { border: 1px solid #ddd; padding: 8px; text-align: left; }
            th { background-color: #f2f2f2; }
            tr:nth-child(even) { background-color: #f9f9f9; }
            h1, h2 { color: #333; }
            .chart-container { margin: 20px 0; text-align: center; }
            .chart-container img { max-width: 100%; height: auto; }
        </style>
        """

        # Prepare report content
        report_content = "# Folder Metrics Report\n\n"
        report_content += (
            f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
        )

        # Sort folders for consistent output (excluding _config)
        sorted_folders = sorted([k for k in folder_data.keys() if k != "_config"])

        if not sorted_folders:
            if log_callback:
                log_callback("처리할 폴더 데이터가 없습니다.\n")
            return None, None, None, None

        report_content += "## Summary by Folder\n\n"
        report_content += "Note: Averages are calculated after excluding the highest and lowest values.\n\n"

        # Summary by Folder - new order
        summary_headers = [
            "Date",
            "Network Type",
            "Carrier",
            "City",
            "Area",
            "Region",
            "Device",
            "Game",
            "Average FPS",
            "Average Bandwidth (Mbps)",
            "Average RTT (ms)",
            "Average Playtime (s)",
            "Folder",
            "Number of Files",
        ]
        report_content += "| " + " | ".join(summary_headers) + " |\n"
        report_content += (
            "|" + "|".join(["-" * len(h) for h in summary_headers]) + "|\n"
        )

        for folder in sorted_folders:
            folder_info = folder_data[folder]
            if not isinstance(folder_info, dict):
                folder_info = dict(folder_info)
            files = list(folder_info.get("files", []))
            num_files = len(files)
            path_components = parse_folder_path(folder)
            # Calculate averages
            valid_playtimes = [
                file.get("playtime")
                for file in files
                if isinstance(file.get("playtime"), (int, float))
                and file.get("playtime") > 0
            ]
            valid_fps = [
                file.get("fps")
                for file in files
                if isinstance(file.get("fps"), (int, float)) and file.get("fps") > 0
            ]
            valid_bandwidths = [
                file.get("bandwidth")
                for file in files
                if isinstance(file.get("bandwidth"), (int, float))
                and file.get("bandwidth") > 0
            ]
            valid_rtts = [
                file.get("rtt")
                for file in files
                if isinstance(file.get("rtt"), (int, float)) and file.get("rtt") > 0
            ]
            # Calculate averages (excluding min/max if enough values)
            if len(valid_playtimes) >= 3:
                sorted_playtimes = sorted(valid_playtimes)
                avg_playtime = sum(sorted_playtimes[1:-1]) / len(sorted_playtimes[1:-1])
            elif valid_playtimes:
                avg_playtime = sum(valid_playtimes) / len(valid_playtimes)
            else:
                avg_playtime = None
            if len(valid_fps) >= 3:
                sorted_fps = sorted(valid_fps)
                avg_fps = sum(sorted_fps[1:-1]) / len(sorted_fps[1:-1])
            elif valid_fps:
                avg_fps = sum(valid_fps) / len(valid_fps)
            else:
                avg_fps = None
            if len(valid_bandwidths) >= 3:
                sorted_bw = sorted(valid_bandwidths)
                avg_bw = sum(sorted_bw[1:-1]) / len(sorted_bw[1:-1])
            elif valid_bandwidths:
                avg_bw = sum(valid_bandwidths) / len(valid_bandwidths)
            else:
                avg_bw = None
            if len(valid_rtts) >= 3:
                sorted_rtt = sorted(valid_rtts)
                avg_rtt = sum(sorted_rtt[1:-1]) / len(sorted_rtt[1:-1])
            elif valid_rtts:
                avg_rtt = sum(valid_rtts) / len(valid_rtts)
            else:
                avg_rtt = None
            rel_folder = os.path.relpath(folder, root_abs)
            if rel_folder == ".":
                rel_folder = "(root)"

            def fmt(val):
                return (
                    f"{val:.2f}"
                    if isinstance(val, (int, float)) and val is not None
                    else (val if val not in [None, -1] else "N/A")
                )

            row_data = {
                "Date": path_components["date"],
                "Network Type": path_components["network"],
                "Carrier": path_components["carrier"],
                "City": path_components["city"],
                "Area": path_components["area"],
                "Region": path_components["region"],
                "Device": path_components["device"],
                "Game": path_components["game"],
                "Average FPS": fmt(avg_fps),
                "Average Bandwidth (Mbps)": fmt(avg_bw),
                "Average RTT (ms)": fmt(avg_rtt),
                "Average Playtime (s)": fmt(avg_playtime),
                "Folder": rel_folder,
                "Number of Files": str(num_files),
            }
            row = [str(row_data[h]) for h in summary_headers]
            report_content += "| " + " | ".join(row) + " |\n"

        # Add detailed metrics for each folder
        report_content += "\n## Detailed Metrics by Folder\n\n"
        detail_headers = [
            "Date",
            "Network Type",
            "Carrier",
            "City",
            "Area",
            "Region",
            "Device",
            "Game",
            "FPS",
            "Bandwidth (Mbps)",
            "RTT (ms)",
            "Playtime (s)",
            "Folder",
            "Filename",
        ]
        for folder in sorted_folders:
            folder_info = folder_data[folder]
            if not isinstance(folder_info, dict):
                folder_info = dict(folder_info)
            files = list(folder_info.get("files", []))
            rel_folder = os.path.relpath(folder, root_abs)
            if rel_folder == ".":
                rel_folder = "(root)"
            report_content += f"### {rel_folder}\n\n"

            # Create table header (new order)
            report_content += "| " + " | ".join(detail_headers) + " |\n"
            report_content += (
                "|" + "|".join(["-" * len(h) for h in detail_headers]) + "|\n"
            )

            # Add file data in new order
            for file_data in sorted(files, key=lambda x: x.get("filename", "")):
                path_components = parse_folder_path(folder)
                row_data = {
                    "Date": path_components["date"],
                    "Network Type": path_components["network"],
                    "Carrier": path_components["carrier"],
                    "City": path_components["city"],
                    "Area": path_components["area"],
                    "Region": path_components["region"],
                    "Device": path_components["device"],
                    "Game": path_components["game"],
                    "FPS": file_data.get("fps", "N/A"),
                    "Bandwidth (Mbps)": file_data.get("bandwidth", "N/A"),
                    "RTT (ms)": file_data.get("rtt", "N/A"),
                    "Playtime (s)": file_data.get("playtime", "N/A"),
                    "Folder": rel_folder,
                    "Filename": file_data.get("filename", "Unknown"),
                }

                # Format values
                def fmt(val):
                    return (
                        f"{val:.2f}"
                        if isinstance(val, (int, float)) and val > 0
                        else (val if val != -1 else "N/A")
                    )

                row = [fmt(row_data[h]) for h in detail_headers]
                report_content += "| " + " | ".join(row) + " |\n"

            report_content += "\n"

        if log_callback:
            log_callback("리포트 파일 저장 중...\n")

        try:
            # Save Markdown reports
            with open(fixed_md, "w", encoding="utf-8") as f:
                f.write(report_content)

            with open(timestamped_md, "w", encoding="utf-8") as f:
                f.write(report_content)

            # Convert to HTML and save
            html_content = markdown.markdown(report_content, extensions=["tables"])
            html_full_content = f"<!DOCTYPE html><html><head>{html_style}</head><body>{html_content}</body></html>"

            with open(fixed_html, "w", encoding="utf-8") as f:
                f.write(html_full_content)

            with open(timestamped_html, "w", encoding="utf-8") as f:
                f.write(html_full_content)

            if log_callback:
                log_callback(
                    f"생성된 파일:\n- {fixed_md}\n- {timestamped_md}\n- {fixed_html}\n- {timestamped_html}\n"
                )

            return fixed_md, timestamped_md, fixed_html, timestamped_html

        except Exception as e:
            if log_callback:
                log_callback(f"파일 저장 중 오류 발생: {e}\n")
            return None, None, None, None

    except Exception as e:
        if log_callback:
            log_callback(f"리포트 생성 중 오류 발생: {str(e)}\n")
        return None, None, None, None


def export_to_excel(folder_data, reports_dir=None, timestamp=None, log_callback=None):
    """Export data to Excel files - one for detailed data and one for averages."""
    if log_callback:
        log_callback("Excel 파일 생성 중...\n")

    try:
        # Get reports directory and root_abs from folder_data if not provided
        if reports_dir is None:
            reports_dir = folder_data.get("_config", {}).get("reports_dir", "reports")
        root_abs = folder_data.get("_config", {}).get("root_abs", "")

        # Create timestamp if not provided
        if timestamp is None:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

        # Create Excel files
        details_filename = os.path.join(
            reports_dir, f"metrics_details_{timestamp}.xlsx"
        )
        averages_filename = os.path.join(
            reports_dir, f"metrics_averages_{timestamp}.xlsx"
        )

        if log_callback:
            log_callback("Excel 워크북 생성 중...\n")

        # Create workbook for detailed data
        wb_details = Workbook()
        ws_details = wb_details.active
        ws_details.title = "Metrics Details"

        # Define alternating colors
        color1 = "FFFFFF"  # White
        color2 = "F5F5F5"  # Very light gray

        # Set up headers (new order)
        headers = [
            "Date",
            "Network Type",
            "Carrier",
            "City",
            "Area",
            "Region",
            "Device",
            "Game",
            "FPS",
            "Bandwidth (Mbps)",
            "RTT (ms)",
            "Playtime (s)",
            "Folder",
            "Filename",
        ]

        # Set up headers for averages (new order)
        avg_headers = [
            "Date",
            "Network Type",
            "Carrier",
            "City",
            "Area",
            "Region",
            "Device",
            "Game",
            "Average FPS",
            "Average Bandwidth (Mbps)",
            "Average RTT (ms)",
            "Average Playtime (s)",
            "Folder",
            "Number of Files",
        ]

        if log_callback:
            log_callback("상세 데이터 시트 작성 중...\n")

        # Set up headers for details worksheet
        for col, header in enumerate(headers, 1):
            cell = ws_details.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Get sorted folders (excluding _config)
        sorted_folders = sorted([k for k in folder_data.keys() if k != "_config"])

        if not sorted_folders:
            if log_callback:
                log_callback("처리할 폴더 데이터가 없습니다.\n")
            return None, None

        # Prepare sorted data
        all_data = []
        for folder in sorted_folders:
            folder_info = folder_data[folder]
            if not isinstance(folder_info, dict):
                folder_info = dict(folder_info)
            files = list(folder_info.get("files", []))

            # Parse folder path
            path_components = parse_folder_path(folder)

            # 폴더 경로를 선택한 폴더(root_abs) 기준 상대경로로 변환
            rel_folder = os.path.relpath(folder, root_abs)
            if rel_folder == ".":
                rel_folder = "(root)"

            for file_data in files:
                data_row = {
                    "Date": path_components["date"],
                    "Network Type": path_components["network"],
                    "Carrier": path_components["carrier"],
                    "City": path_components["city"],
                    "Area": path_components["area"],
                    "Region": path_components["region"],
                    "Device": path_components["device"],
                    "Game": path_components["game"],
                    "FPS": file_data.get("fps", -1),
                    "Bandwidth (Mbps)": file_data.get("bandwidth", -1),
                    "RTT (ms)": file_data.get("rtt", -1),
                    "Playtime (s)": file_data.get("playtime", -1),
                    "Folder": rel_folder,
                    "Filename": file_data.get("filename", "Unknown"),
                }
                all_data.append(data_row)

        # Sort data by folder name
        all_data.sort(key=lambda x: x["Folder"].lower())

        # Fill details data with alternating colors
        row = 2
        current_folder = None
        color_toggle = True  # True for color1, False for color2

        for data in all_data:
            # Check if folder changed
            if current_folder != data["Folder"]:
                current_folder = data["Folder"]
                color_toggle = not color_toggle  # Switch color for new folder

            # Apply background color to entire row
            fill_color = color1 if color_toggle else color2
            for col in range(1, len(headers) + 1):
                cell = ws_details.cell(row=row, column=col)
                cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            # Fill data in new order
            for col, header in enumerate(headers, 1):
                ws_details.cell(row=row, column=col, value=data[header])
            row += 1

        # Auto-adjust column widths
        for col in range(1, len(headers) + 1):
            ws_details.column_dimensions[get_column_letter(col)].auto_size = True

        if log_callback:
            log_callback("평균값 시트 작성 중...\n")

        # Create workbook for averages
        wb_averages = Workbook()
        ws_averages = wb_averages.active
        ws_averages.title = "Folder Averages"

        # Set up headers for averages worksheet
        for col, header in enumerate(avg_headers, 1):
            cell = ws_averages.cell(row=1, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(
                start_color="CCCCCC", end_color="CCCCCC", fill_type="solid"
            )
            cell.alignment = Alignment(horizontal="center")

        # Calculate and fill averages data with alternating colors
        row = 2
        color_toggle = True  # Reset color toggle for averages sheet

        for folder in sorted_folders:
            # 폴더 경로를 선택한 폴더(root_abs) 기준 상대경로로 변환
            rel_folder = os.path.relpath(folder, root_abs)
            if rel_folder == ".":
                rel_folder = "(root)"

            folder_info = folder_data[folder]
            if not isinstance(folder_info, dict):
                folder_info = dict(folder_info)
            files = list(folder_info.get("files", []))
            path_components = parse_folder_path(folder)

            # Calculate averages
            fps_values = [f.get("fps", -1) for f in files if f.get("fps", -1) > 0]
            bw_values = [
                f.get("bandwidth", -1) for f in files if f.get("bandwidth", -1) > 0
            ]
            rtt_values = [f.get("rtt", -1) for f in files if f.get("rtt", -1) > 0]
            playtime_values = [
                f.get("playtime", -1) for f in files if f.get("playtime", -1) > 0
            ]

            # Calculate averages (excluding min/max if enough values)
            if len(fps_values) >= 3:
                fps_values = sorted(fps_values)[1:-1]
            if len(bw_values) >= 3:
                bw_values = sorted(bw_values)[1:-1]
            if len(rtt_values) >= 3:
                rtt_values = sorted(rtt_values)[1:-1]
            if len(playtime_values) >= 3:
                playtime_values = sorted(playtime_values)[1:-1]

            avg_fps = sum(fps_values) / len(fps_values) if fps_values else -1
            avg_bw = sum(bw_values) / len(bw_values) if bw_values else -1
            avg_rtt = sum(rtt_values) / len(rtt_values) if rtt_values else -1
            avg_playtime = (
                sum(playtime_values) / len(playtime_values) if playtime_values else -1
            )

            # Apply background color to entire row
            fill_color = color1 if color_toggle else color2
            for col in range(1, len(avg_headers) + 1):
                cell = ws_averages.cell(row=row, column=col)
                cell.fill = PatternFill(
                    start_color=fill_color, end_color=fill_color, fill_type="solid"
                )

            # Fill data for averages in new order
            avg_data = {
                "Date": path_components["date"],
                "Network Type": path_components["network"],
                "Carrier": path_components["carrier"],
                "City": path_components["city"],
                "Area": path_components["area"],
                "Region": path_components["region"],
                "Device": path_components["device"],
                "Game": path_components["game"],
                "Average FPS": round(avg_fps, 2) if avg_fps > 0 else "N/A",
                "Average Bandwidth (Mbps)": round(avg_bw, 2) if avg_bw > 0 else "N/A",
                "Average RTT (ms)": round(avg_rtt, 2) if avg_rtt > 0 else "N/A",
                "Average Playtime (s)": (
                    round(avg_playtime, 2) if avg_playtime > 0 else "N/A"
                ),
                "Folder": rel_folder,
                "Number of Files": len(files),
            }
            for col, header in enumerate(avg_headers, 1):
                ws_averages.cell(row=row, column=col, value=avg_data[header])

            row += 1
            color_toggle = not color_toggle  # Switch color for next folder

        # Auto-adjust column widths for averages
        for col in range(1, len(avg_headers) + 1):
            ws_averages.column_dimensions[get_column_letter(col)].auto_size = True

        if log_callback:
            log_callback("Excel 파일 저장 중...\n")

        # Save both workbooks
        wb_details.save(details_filename)
        wb_averages.save(averages_filename)

        if log_callback:
            log_callback(f"생성된 파일:\n- {details_filename}\n- {averages_filename}\n")

        # CSV 파일로도 저장 (상세/평균)
        try:
            # 상세 데이터 CSV
            details_csv = os.path.join(reports_dir, f"metrics_details_{timestamp}.csv")
            df_details = pd.DataFrame(all_data, columns=headers)
            df_details.to_csv(details_csv, index=False, encoding="utf-8-sig")
            # 평균 데이터 CSV
            averages_csv = os.path.join(
                reports_dir, f"metrics_averages_{timestamp}.csv"
            )
            avg_rows = []
            for folder in sorted_folders:
                rel_folder = os.path.relpath(folder, root_abs)
                if rel_folder == ".":
                    rel_folder = "(root)"
                folder_info = folder_data[folder]
                if not isinstance(folder_info, dict):
                    folder_info = dict(folder_info)
                files = list(folder_info.get("files", []))
                path_components = parse_folder_path(folder)
                fps_values = [f.get("fps", -1) for f in files if f.get("fps", -1) > 0]
                bw_values = [
                    f.get("bandwidth", -1) for f in files if f.get("bandwidth", -1) > 0
                ]
                rtt_values = [f.get("rtt", -1) for f in files if f.get("rtt", -1) > 0]
                playtime_values = [
                    f.get("playtime", -1) for f in files if f.get("playtime", -1) > 0
                ]
                if len(fps_values) >= 3:
                    fps_values = sorted(fps_values)[1:-1]
                if len(bw_values) >= 3:
                    bw_values = sorted(bw_values)[1:-1]
                if len(rtt_values) >= 3:
                    rtt_values = sorted(rtt_values)[1:-1]
                if len(playtime_values) >= 3:
                    playtime_values = sorted(playtime_values)[1:-1]
                avg_fps = sum(fps_values) / len(fps_values) if fps_values else -1
                avg_bw = sum(bw_values) / len(bw_values) if bw_values else -1
                avg_rtt = sum(rtt_values) / len(rtt_values) if rtt_values else -1
                avg_playtime = (
                    sum(playtime_values) / len(playtime_values)
                    if playtime_values
                    else -1
                )
                avg_data = {
                    "Date": path_components["date"],
                    "Network Type": path_components["network"],
                    "Carrier": path_components["carrier"],
                    "City": path_components["city"],
                    "Area": path_components["area"],
                    "Region": path_components["region"],
                    "Device": path_components["device"],
                    "Game": path_components["game"],
                    "Average FPS": round(avg_fps, 2) if avg_fps > 0 else "N/A",
                    "Average Bandwidth (Mbps)": (
                        round(avg_bw, 2) if avg_bw > 0 else "N/A"
                    ),
                    "Average RTT (ms)": round(avg_rtt, 2) if avg_rtt > 0 else "N/A",
                    "Average Playtime (s)": (
                        round(avg_playtime, 2) if avg_playtime > 0 else "N/A"
                    ),
                    "Folder": rel_folder,
                    "Number of Files": len(files),
                }
                avg_rows.append(avg_data)
            df_avg = pd.DataFrame(avg_rows, columns=avg_headers)
            df_avg.to_csv(averages_csv, index=False, encoding="utf-8-sig")
            if log_callback:
                log_callback(f"CSV 파일도 생성됨:\n- {details_csv}\n- {averages_csv}\n")
        except Exception as e:
            if log_callback:
                log_callback(f"CSV 저장 중 오류 발생: {e}\n")

        return details_filename, averages_filename

    except Exception as e:
        if log_callback:
            log_callback(f"Excel 파일 생성 중 오류 발생: {e}\n")
            log_callback(traceback.format_exc())  # <-- 전체 스택 트레이스 출력
        return None, None


def generate_performance_plots(folder_data, timestamp, log_callback=None):
    """Generate performance visualization plots grouped by Region and Carrier."""
    if log_callback:
        log_callback("성능 차트 생성 중...\n")

    try:
        # Create plots directory if it doesn't exist
        plots_dir = os.path.join("reports", "plots")
        os.makedirs(plots_dir, exist_ok=True)

        if log_callback:
            log_callback("데이터 준비 중...\n")

        # Prepare data for plotting with region and carrier information
        plot_data = []
        for folder, data in folder_data.items():
            fps_values = []
            bandwidth_values = []
            rtt_values = []

            # Parse folder path to get region and carrier
            path_components = parse_folder_path(folder)
            region = path_components["region"]
            carrier = path_components["carrier"]

            for file_data in data["files"]:
                if isinstance(file_data, dict):
                    if (
                        "fps" in file_data
                        and isinstance(file_data["fps"], (int, float))
                        and file_data["fps"] > -1
                    ):
                        fps_values.append(file_data["fps"])
                    if (
                        "bandwidth" in file_data
                        and isinstance(file_data["bandwidth"], (int, float))
                        and file_data["bandwidth"] > -1
                    ):
                        bandwidth_values.append(file_data["bandwidth"])
                    if (
                        "rtt" in file_data
                        and isinstance(file_data["rtt"], (int, float))
                        and file_data["rtt"] > -1
                    ):
                        rtt_values.append(file_data["rtt"])

            if fps_values or bandwidth_values or rtt_values:
                plot_data.append(
                    {
                        "folder": folder,
                        "region": region,
                        "carrier": carrier,
                        "avg_fps": float(np.mean(fps_values)) if fps_values else 0.0,
                        "avg_bandwidth": (
                            float(np.mean(bandwidth_values))
                            if bandwidth_values
                            else 0.0
                        ),
                        "avg_rtt": float(np.mean(rtt_values)) if rtt_values else 0.0,
                        "fps_values": fps_values,
                        "bandwidth_values": bandwidth_values,
                        "rtt_values": rtt_values,
                    }
                )

        if not plot_data:
            if log_callback:
                log_callback("차트를 생성할 데이터가 없습니다.\n")
            return False

        # Convert to DataFrame for plotting
        df = pd.DataFrame(plot_data)

        if log_callback:
            log_callback("차트 생성 시작...\n")

        # Set figure size and font size for better readability
        plt.rcParams.update(
            {"font.size": 10, "figure.autolayout": True, "figure.figsize": (20, 10)}
        )

        # Group data by region and carrier
        regions = sorted(df["region"].unique())
        carriers = sorted(df["carrier"].unique())

        # 1. Bar plots for averages by region
        for region in regions:
            if log_callback:
                log_callback(f"리전별 평균 차트 생성 중: {region}\n")

            region_data = df[df["region"] == region]
            if not region_data.empty:
                fig, ax = plt.subplots()
                x = np.arange(len(region_data))
                width = 0.25

                ax.bar(
                    x - width,
                    region_data["avg_fps"],
                    width,
                    label="FPS",
                    color="skyblue",
                )
                ax.bar(
                    x,
                    region_data["avg_bandwidth"],
                    width,
                    label="Bandwidth (Mbps)",
                    color="lightgreen",
                )
                ax.bar(
                    x + width,
                    region_data["avg_rtt"],
                    width,
                    label="RTT (ms)",
                    color="salmon",
                )

                ax.set_title(f"Average Performance Metrics - Region: {region}")
                ax.set_xlabel("Folder")
                ax.set_ylabel("Value")
                ax.set_xticks(x)
                ax.set_xticklabels(region_data["folder"], rotation=45, ha="right")
                ax.legend()
                ax.grid(True, linestyle="--", alpha=0.3)

                plt.tight_layout()
                plt.savefig(
                    os.path.join(
                        plots_dir, f"averages_bar_plot_region_{region}_{timestamp}.png"
                    ),
                    dpi=300,
                    bbox_inches="tight",
                )
                plt.close()

        # 2. Bar plots for averages by carrier
        for carrier in carriers:
            if log_callback:
                log_callback(f"통신사별 평균 차트 생성 중: {carrier}\n")

            carrier_data = df[df["carrier"] == carrier]
            if not carrier_data.empty:
                fig, ax = plt.subplots()
                x = np.arange(len(carrier_data))
                width = 0.25

                ax.bar(
                    x - width,
                    carrier_data["avg_fps"],
                    width,
                    label="FPS",
                    color="skyblue",
                )
                ax.bar(
                    x,
                    carrier_data["avg_bandwidth"],
                    width,
                    label="Bandwidth (Mbps)",
                    color="lightgreen",
                )
                ax.bar(
                    x + width,
                    carrier_data["avg_rtt"],
                    width,
                    label="RTT (ms)",
                    color="salmon",
                )

                ax.set_title(f"Average Performance Metrics - Carrier: {carrier}")
                ax.set_xlabel("Folder")
                ax.set_ylabel("Value")
                ax.set_xticks(x)
                ax.set_xticklabels(carrier_data["folder"], rotation=45, ha="right")
                ax.legend()
                ax.grid(True, linestyle="--", alpha=0.3)

                plt.tight_layout()
                plt.savefig(
                    os.path.join(
                        plots_dir,
                        f"averages_bar_plot_carrier_{carrier}_{timestamp}.png",
                    ),
                    dpi=300,
                    bbox_inches="tight",
                )
                plt.close()

        # 3. Box plots for distributions
        metrics = ["fps_values", "bandwidth_values", "rtt_values"]
        titles = [
            "FPS Distribution",
            "Bandwidth Distribution (Mbps)",
            "RTT Distribution (ms)",
        ]
        colors = ["skyblue", "lightgreen", "salmon"]

        # Box plots by region
        for region in regions:
            if log_callback:
                log_callback(f"리전별 분포도 차트 생성 중: {region}\n")

            region_data = df[df["region"] == region]
            if not region_data.empty:
                for metric, title, color in zip(metrics, titles, colors):
                    fig, ax = plt.subplots()
                    data_to_plot = [
                        d[metric] for _, d in region_data.iterrows() if d[metric]
                    ]
                    if data_to_plot:
                        bp = ax.boxplot(data_to_plot, patch_artist=True)
                        for element in [
                            "boxes",
                            "whiskers",
                            "fliers",
                            "means",
                            "medians",
                            "caps",
                        ]:
                            plt.setp(bp[element], color="black")
                        plt.setp(bp["boxes"], facecolor=color)

                        ax.set_title(f"{title} - Region: {region}")
                        ax.set_xticks(range(1, len(data_to_plot) + 1))
                        ax.set_xticklabels(
                            [
                                d["folder"]
                                for _, d in region_data.iterrows()
                                if d[metric]
                            ],
                            rotation=45,
                            ha="right",
                        )
                        ax.grid(True, linestyle="--", alpha=0.3)

                        plt.tight_layout()
                        metric_name = metric.split("_")[0]
                        plt.savefig(
                            os.path.join(
                                plots_dir,
                                f"{metric_name}_boxplot_region_{region}_{timestamp}.png",
                            ),
                            dpi=300,
                            bbox_inches="tight",
                        )
                        plt.close()

        # Box plots by carrier
        for carrier in carriers:
            if log_callback:
                log_callback(f"통신사별 분포도 차트 생성 중: {carrier}\n")

            carrier_data = df[df["carrier"] == carrier]
            if not carrier_data.empty:
                for metric, title, color in zip(metrics, titles, colors):
                    fig, ax = plt.subplots()
                    data_to_plot = [
                        d[metric] for _, d in carrier_data.iterrows() if d[metric]
                    ]
                    if data_to_plot:
                        bp = ax.boxplot(data_to_plot, patch_artist=True)
                        for element in [
                            "boxes",
                            "whiskers",
                            "fliers",
                            "means",
                            "medians",
                            "caps",
                        ]:
                            plt.setp(bp[element], color="black")
                        plt.setp(bp["boxes"], facecolor=color)

                        ax.set_title(f"{title} - Carrier: {carrier}")
                        ax.set_xticks(range(1, len(data_to_plot) + 1))
                        ax.set_xticklabels(
                            [
                                d["folder"]
                                for _, d in carrier_data.iterrows()
                                if d[metric]
                            ],
                            rotation=45,
                            ha="right",
                        )
                        ax.grid(True, linestyle="--", alpha=0.3)
                        plt.tight_layout()
                        metric_name = metric.split("_")[0]
                        plt.savefig(
                            os.path.join(
                                plots_dir,
                                f"{metric_name}_boxplot_carrier_{carrier}_{timestamp}.png",
                            ),
                            dpi=300,
                            bbox_inches="tight",
                        )
                        plt.close()

        if log_callback:
            log_callback(f"모든 차트가 {plots_dir}/ 폴더에 저장되었습니다.\n")

        return True

    except Exception as e:
        if log_callback:
            log_callback(f"차트 생성 중 오류 발생: {e}\n")
        return False


def parse_folder_path(path):
    """Parse folder path into components."""
    # Split the path and get all parts
    parts = path.split(os.sep)

    # Initialize result with default values
    result = {
        "date": "Unknown",
        "city": "Unknown",
        "area": "Unknown",
        "region": "Unknown",
        "carrier": "Unknown",
        "network": "Unknown",
        "game": "Unknown",
        "device": "Unknown",
    }

    # Find the date component (YYYYMMDD format)
    date_pattern = re.compile(r"^\d{8}$")
    date_index = -1
    for i, part in enumerate(parts):
        if date_pattern.match(part):
            try:
                year = int(part[:4])
                month = int(part[4:6])
                day = int(part[6:8])
                if 2000 <= year <= 2100 and 1 <= month <= 12 and 1 <= day <= 31:
                    result["date"] = part
                    date_index = i
                    break
            except ValueError:
                continue

    if date_index != -1:
        # Get the components after the date
        remaining_parts = parts[date_index + 1 :]

        # Map the remaining components based on the carrier type
        if len(remaining_parts) >= 4:
            result["city"] = remaining_parts[0]
            result["area"] = remaining_parts[1]
            result["region"] = remaining_parts[2]
            result["carrier"] = remaining_parts[3]

            # Check if carrier is wifi (case-insensitive)
            if remaining_parts[3].lower() == "wifi":
                # For WiFi, network is set to "WIFI" and remaining components shift
                result["carrier"] = ""  # wifi인 경우 carrier를 빈 문자열로
                result["network"] = "WIFI"
                if len(remaining_parts) >= 5:
                    result["game"] = remaining_parts[4]
                if len(remaining_parts) >= 6:
                    result["device"] = remaining_parts[5]
            else:
                # For other carriers (e.g., voda), expect network type
                if len(remaining_parts) >= 5:
                    network = remaining_parts[4].lower()
                    if network in ["4g", "5g"]:
                        result["network"] = network.upper()
                    else:
                        result["network"] = ""
                if len(remaining_parts) >= 6:
                    result["game"] = remaining_parts[5]
                if len(remaining_parts) >= 7:
                    result["device"] = remaining_parts[6]

    # Clean up any empty or whitespace-only values
    for key in list(result.keys()):
        if not result[key] or result[key].strip() == "":
            result[key] = ""
        else:
            result[key] = result[key].strip()

    return result


def main():
    """Main function to process PDF files and generate folder-based report."""
    parser = argparse.ArgumentParser(
        description="Process PDF files and generate reports"
    )
    parser.add_argument("--excel", action="store_true", help="Generate Excel reports")
    parser.add_argument(
        "--plots", action="store_true", help="Generate performance plots"
    )
    parser.add_argument("--folder", help="Specify folder path directly")
    args = parser.parse_args()

    def console_log(message):
        """Print log messages to console."""
        print(message, end="")
        sys.stdout.flush()

    try:
        # 폴더 경로가 지정된 경우에만 처리
        if not args.folder:
            console_log("폴더 경로를 지정해주세요.\n")
            sys.exit(1)

        folder_path = args.folder
        console_log(f"선택한 폴더: {folder_path}\n")

        # Create reports and plots directories
        reports_dir = os.path.join(os.path.dirname(folder_path), "reports")
        plots_dir = os.path.join(reports_dir, "plots")
        os.makedirs(reports_dir, exist_ok=True)
        os.makedirs(plots_dir, exist_ok=True)
        console_log(f"생성된 디렉토리:\n- {reports_dir}\n- {plots_dir}\n")

        # Process PDF files
        console_log("\nPDF 파일 처리 시작...\n")
        folder_data = process_pdf_files(folder_path, log_callback=console_log)

        if folder_data:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")

            # Generate markdown and HTML reports
            console_log("\n마크다운 및 HTML 리포트 생성 중...\n")
            fixed_md, timestamped_md, fixed_html, timestamped_html = (
                generate_folder_report(folder_data, log_callback=console_log)
            )

            # Generate Excel reports if requested
            if args.excel:
                console_log("\nExcel 리포트 생성 중...\n")
                details_excel, averages_excel = export_to_excel(
                    folder_data, reports_dir, timestamp, log_callback=console_log
                )
                if details_excel and averages_excel:
                    console_log(f"\nExcel 리포트 생성됨:\n")
                    console_log(f"1. 상세 리포트: {details_excel}\n")
                    console_log(f"2. 평균값 리포트: {averages_excel}\n")
                else:
                    console_log("\nExcel 리포트 생성 실패.\n")

            # Generate plots if requested
            if args.plots:
                console_log("\n성능 차트 생성 중...\n")
                plots_success = generate_performance_plots(
                    folder_data, timestamp, log_callback=console_log
                )
                if plots_success:
                    console_log(f"\n성능 차트가 {plots_dir} 폴더에 저장되었습니다.\n")
                else:
                    console_log("\n성능 차트 생성 실패.\n")

            if fixed_md and timestamped_md and fixed_html and timestamped_html:
                console_log(f"\n생성된 리포트:\n")
                console_log(
                    f"1. 고정 파일명 마크다운 리포트: {os.path.abspath(fixed_md)}\n"
                )
                console_log(
                    f"2. 타임스탬프 마크다운 리포트: {os.path.abspath(timestamped_md)}\n"
                )
                console_log(
                    f"3. 고정 파일명 HTML 리포트: {os.path.abspath(fixed_html)}\n"
                )
                console_log(
                    f"4. 타임스탬프 HTML 리포트: {os.path.abspath(timestamped_html)}\n"
                )
            else:
                console_log("\n리포트 생성 실패.\n")
        else:
            console_log("\n데이터 처리 실패. 리포트가 생성되지 않았습니다.\n")

    except Exception as e:
        console_log(f"메인 함수 실행 중 오류 발생: {e}\n")
        sys.exit(1)


if __name__ == "__main__":
    main()
